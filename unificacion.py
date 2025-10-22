    
import pandas as pd
import os
from sqlalchemy import create_engine
import math
import sys
from pathlib import Path
import re
import time
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook

# ======================================================================================
# FUNCIONES AUXILIARES PARA MANEJO DE RUTAS (EXE o Script)
# ======================================================================================

def obtener_ruta_salida():
    """
    Obtiene la ruta donde guardar archivos (CSV, Excel)
    - Si se ejecuta como .exe: usa la carpeta del ejecutable
    - Si se ejecuta como .py: usa la carpeta actual o Descargas si no tiene permisos
    
    Prioridad:
    1. Carpeta del ejecutable/script
    2. Carpeta Descargas del usuario
    """
    try:
        # Si es ejecutable compilado
        if getattr(sys, 'frozen', False):
            ruta_exe = Path(sys.executable).parent
            return str(ruta_exe)
        # Si es script de Python
        else:
            ruta_script = Path(__file__).parent
            # Validar si se puede escribir en la carpeta del script
            if os.access(ruta_script, os.W_OK):
                return str(ruta_script)
    except:
        pass
    
    # Plan B: Descargas
    try:
        from pathlib import Path
        descargas = Path.home() / "Downloads"
        if descargas.exists():
            return str(descargas)
    except:
        pass
    
    # Plan C: Temp
    try:
        import tempfile
        return tempfile.gettempdir()
    except:
        pass
    
    # Plan D: Carpeta actual
    return os.getcwd()

# Obtener ruta de salida al iniciar
RUTA_SALIDA = obtener_ruta_salida()
print(f"[INFO] Archivos se guardarán en: {RUTA_SALIDA}", flush=True)

# ======================================================================================
# FUNCIÓN: EXTRAER ANCHO DE ROLLO DESDE NOMBRE_COMPONENTE
# ======================================================================================

def extraer_ancho_rollo_mm(nombre_componente: str) -> float:
    """
    Extrae el ancho del rollo en mm desde el nombre del componente.
    
    Busca patrones como:
    - "280MM"
    - "280 MM"
    - "280mm"
    - "280 mm"
    
    Retorna el número encontrado o None si no hay coincidencia.
    
    Args:
        nombre_componente: Nombre del componente (ej: "LAMINADO FILM PETG 280 MM 45 MIC")
    
    Returns:
        float: Ancho del rollo en mm, o None si no encuentra
    """
    if not nombre_componente or not isinstance(nombre_componente, str):
        return None
    
    # Expresión regular para encontrar número seguido de espacios opcionales y MM
    # Busca: número + espacios opcionales + MM (case insensitive)
    patron = r'(\d+)\s*MM(?:\s|$)'
    
    # Buscar todas las coincidencias
    coincidencias = re.findall(patron, nombre_componente, re.IGNORECASE)
    
    if coincidencias:
        # Tomar la primera coincidencia y convertir a float
        try:
            return float(coincidencias[0])
        except (ValueError, IndexError):
            return None
    
    return None

# ======================================================================================
# MOTOR DE CÁLCULO - FUNCIONES VITALES INTEGRADAS
# ======================================================================================

FACTOR_CILINDRO_A_MM = 3.175  # Z a milímetros
ROLLO_BASE_MM = 330  # Rollo estándar 330 mm
BANDERA_HORIZONTAL = 20  # mm de bandera (se restan del ancho útil)
GAP_HORIZONTAL_OBJETIVO = 2.7  # Gap objetivo
GAP_HORIZONTAL_MIN = 2.3
GAP_HORIZONTAL_MAX = 20
GAP_VERTICAL_OBJETIVO = 2.7
GAP_VERTICAL_MIN = 2.5
GAP_VERTICAL_MAX = 20
MAX_REPETICIONES_VERTICALES = 8
AJUSTE_DESARROLLO_REP_MM = 0.75  # mm adicionales por repeat

# Conjuntos de cilindros disponibles
CILINDROS_FB = sorted([60, 67, 70, 74, 77, 80, 84, 88, 91, 97, 99, 102, 105, 107, 108, 111,
                        116, 117, 122, 127, 129, 168], reverse=True)

class EvaluacionZUniforme:
    """Resultado de evaluación de un Z particular"""
    def __init__(self, z, desarrollo_mm, n, gap, diff_obj, valido, detalle):
        self.z = z
        self.desarrollo_mm = desarrollo_mm
        self.n = n
        self.gap = gap
        self.diff_obj = diff_obj
        self.valido = valido
        self.detalle = detalle

def evaluar_z_uniforme(z: int, alto_etq: float, gap_obj: float,
                       gap_min: float, gap_max: float, max_n: int) -> EvaluacionZUniforme:
    """
    Evalúa un Z particular para determinar cuántas etiquetas caben verticalmente.
    
    Args:
        z: Número de cilindro (60-168)
        alto_etq: Alto de etiqueta en mm
        gap_obj: Gap objetivo en mm (2.7)
        gap_min: Gap mínimo permitido (2.5)
        gap_max: Gap máximo permitido (20)
        max_n: Máximo número de repeticiones verticales
    
    Returns:
        EvaluacionZUniforme con n (repeticiones), gap, y validez
    """
    desarrollo = z * FACTOR_CILINDRO_A_MM
    
    # Si el desarrollo es menor que una etiqueta, no es válido
    if desarrollo < alto_etq:
        return EvaluacionZUniforme(z, desarrollo, 0, 0.0, 9999.0, False, 
                                  f"Desarrollo {desarrollo:.1f} mm < Alto etiqueta {alto_etq} mm")
    
    mejor = None
    
    # Iterar sobre posibles n (repeticiones verticales)
    for n in range(1, max_n + 1):
        if desarrollo < n * alto_etq:
            break
        
        gap = (desarrollo - n * alto_etq) / n
        
        if gap < 0:
            continue
        
        # Para n > 1, validar que gap esté dentro de rango; n=1 siempre es válido
        if n > 1 and (gap < gap_min or gap > gap_max):
            continue
        
        # n=1 es válido incluso con gap grande, pero asignar gap_obj como referencia
        if n == 1:
            gap = max(gap, gap_min)  # Al menos el gap mínimo como referencia
        
        diff = abs(gap - gap_obj)
        cand = (n, gap, diff)
        
        # Seleccionar mejor: Máximo n, luego menor diff del objetivo, luego menor gap
        if mejor is None:
            mejor = cand
        else:
            n_best, gap_best, diff_best = mejor
            if (cand[0] > n_best) or (cand[0] == n_best and (cand[2] < diff_best or 
                (cand[2] == diff_best and cand[1] < gap_best))):
                mejor = cand
    
    if mejor:
        n_sel, gap_sel, diff_sel = mejor
        return EvaluacionZUniforme(z, desarrollo, n_sel, gap_sel, diff_sel, True, "OK")
    
    return EvaluacionZUniforme(z, desarrollo, 0, 0.0, 9999.0, False, "Sin combinación válida de n")

def evaluar_rollo_estandar(ancho_mm: float, ancho_etq: float, gap_h_usado: float, 
                           bandera_mm: float) -> dict:
    """
    Evalúa cómo se distribuyen horizontalmente las etiquetas en el rollo.
    
    Args:
        ancho_mm: Ancho del rollo en mm
        ancho_etq: Ancho de etiqueta en mm
        gap_h_usado: Gap horizontal usado en mm
        bandera_mm: Ancho de bandera en mm (se resta del ancho útil)
    
    Returns:
        Dict con etq_eje (cantidad horizontal), gap_entre, aprovechamiento, etc.
    """
    if ancho_mm <= 0:
        return {"etq_eje": 0, "gap_entre": None, "aprovechamiento_pct": 0.0}
    
    paso = ancho_etq + gap_h_usado
    if paso <= 0:
        return {"etq_eje": 0, "gap_entre": None, "aprovechamiento_pct": 0.0}
    
    # Intentar primero CON bandera (20mm de margen)
    ancho_util = max(0.0, ancho_mm - bandera_mm)
    etq_eje = max(0, int(ancho_util // paso))
    
    # Si no cabe ni una etiqueta CON bandera, ignorar bandera y usar todo el ancho
    if etq_eje <= 0:
        ancho_util = ancho_mm  # Usar ancho completo sin restar bandera
        etq_eje = max(0, int(ancho_util // paso))
    
    # Si aún no cabe ni una etiqueta con paso estándar, permitir al menos 1 etiqueta
    # (el gap será mayor pero la producción sigue siendo posible)
    if etq_eje <= 0 and ancho_util >= ancho_etq:
        etq_eje = 1
        # Recalcular gap con solo 1 etiqueta
        gap_total = max(0.0, ancho_util - ancho_etq)
        gap_entre = gap_total  # Un solo gap a un lado
    else:
        gap_total = max(0.0, ancho_util - etq_eje * ancho_etq)
        gap_entre = gap_total / (etq_eje - 1) if etq_eje > 1 else None
    
    aprovechamiento = (etq_eje * ancho_etq) / ancho_util * 100 if ancho_util > 0 else 0.0
    
    return {
        "etq_eje": etq_eje,
        "gap_entre": gap_entre,
        "aprovechamiento_pct": aprovechamiento,
        "ancho_util": ancho_util
    }

def calcular_metraje_material(unidades: int, etq_repeat: int, desarrollo_mm: float,
                              ancho_rollo_mm: float, ajuste_mm: float) -> dict:
    """
    Calcula metros lineales y metros cuadrados necesarios.
    
    Args:
        unidades: Cantidad de etiquetas a producir
        etq_repeat: Etiquetas por repetición (n_vertical × etq_eje)
        desarrollo_mm: Desarrollo del cilindro en mm
        ancho_rollo_mm: Ancho del rollo en mm
        ajuste_mm: Ajuste adicional por repetición (típicamente 0.75 mm)
    
    Returns:
        Dict con repeticiones, ML, M2
    """
    if etq_repeat <= 0 or desarrollo_mm <= 0 or ancho_rollo_mm <= 0:
        return {"repeticiones_reales": 0.0, "ml": 0.0, "m2": 0.0}
    
    # Calcular numero de vueltas necesarias del cilindro
    repeticiones_reales = unidades / etq_repeat
    
    # Metraje lineal = (repeticiones × desarrollo del cilindro) + merma
    longitud_sin_merma_mm = repeticiones_reales * desarrollo_mm
    merma_mm = repeticiones_reales * ajuste_mm
    longitud_total_mm = longitud_sin_merma_mm + merma_mm
    
    # Convertir a metros
    ml = longitud_total_mm / 1000.0
    
    # Metros cuadrados = ML × ancho del rollo en metros
    m2 = ml * (ancho_rollo_mm / 1000.0)
    
    return {
        "repeticiones_reales": repeticiones_reales,
        "ml": ml,
        "m2": m2
    }

def obtener_z_sugerido(alto_etq: float, ancho_etq: float, unidades: int, ancho_rollo_mm: float = None) -> tuple:
    """
    Busca el Z (cilindro) óptimo que minimiza metraje.
    
    Args:
        alto_etq: Alto de etiqueta en mm
        ancho_etq: Ancho de etiqueta en mm
        unidades: Cantidad de unidades a producir
        ancho_rollo_mm: Ancho del rollo en mm (extrae de Nombre_Componente o usa ROLLO_BASE_MM)
    
    Returns:
        Dict con mejor Z y sus métricas, o dict con rollo/Z pero motivo_imposibilidad si es técnicamente imposible
    """
    # Si no se proporciona ancho_rollo_mm, usar valor por defecto
    if ancho_rollo_mm is None:
        ancho_rollo_mm = ROLLO_BASE_MM
    
    mejor = None
    motivo_imposibilidad = None
    z_fallback = None  # Para guardar Z si es necesario
    
    for z in CILINDROS_FB:
        # Evaluar Z verticalmente
        ev_z = evaluar_z_uniforme(z, alto_etq, GAP_VERTICAL_OBJETIVO,
                                  GAP_VERTICAL_MIN, GAP_VERTICAL_MAX, 
                                  MAX_REPETICIONES_VERTICALES)
        
        if not ev_z.valido:
            # Si el alto no cabe en ningun Z, guardar motivo
            if motivo_imposibilidad is None:
                motivo_imposibilidad = f"Alto {int(alto_etq)}mm no cabe en ningun cilindro disponible (60-168mm)"
            continue
        
        # Evaluar rollo horizontalmente (con ancho_rollo_mm extraído)
        ev_rollo = evaluar_rollo_estandar(ancho_rollo_mm, ancho_etq, 
                                          GAP_HORIZONTAL_OBJETIVO, BANDERA_HORIZONTAL)
        
        if ev_rollo["etq_eje"] <= 0:
            # Guardar este Z como fallback por si no hay ninguno viable
            if z_fallback is None:
                z_fallback = z
                motivo_imposibilidad = f"Ancho {int(ancho_etq)}mm no cabe en rollo {int(ancho_rollo_mm)}mm"
            continue
        
        # Calcular etiquetas por repetición
        etq_repeat = ev_z.n * ev_rollo["etq_eje"]
        
        # Calcular metraje (usando ancho_rollo_mm extraído)
        metraje = calcular_metraje_material(unidades, etq_repeat, ev_z.desarrollo_mm,
                                           ancho_rollo_mm, AJUSTE_DESARROLLO_REP_MM)
        
        # Seleccionar mejor por menor ML
        if mejor is None or metraje["ml"] < mejor["ml"]:
            mejor = {
                "z": z,
                "n": ev_z.n,
                "desarrollo_mm": ev_z.desarrollo_mm,
                "gap_vertical": ev_z.gap,
                "etq_eje": ev_rollo["etq_eje"],
                "gap_horizontal_real": ev_rollo["gap_entre"],
                "etq_repeat": etq_repeat,
                "repeticiones": metraje["repeticiones_reales"],
                "ml": metraje["ml"],
                "m2": metraje["m2"],
                "ancho_rollo_mm": ancho_rollo_mm,
                "es_valido": True
            }
    
    # Si hay solución válida, retornarla
    if mejor is not None:
        return mejor
    
    # Si no hay solución con el ancho especificado y no es el rollo base, intentar con rollo más grande
    if ancho_rollo_mm != ROLLO_BASE_MM:
        z_fallback_330 = None
        motivo_imposibilidad_330 = None
        
        for z in CILINDROS_FB:
            ev_z = evaluar_z_uniforme(z, alto_etq, GAP_VERTICAL_OBJETIVO,
                                      GAP_VERTICAL_MIN, GAP_VERTICAL_MAX, 
                                      MAX_REPETICIONES_VERTICALES)
            if not ev_z.valido:
                continue
            
            ev_rollo = evaluar_rollo_estandar(ROLLO_BASE_MM, ancho_etq, 
                                              GAP_HORIZONTAL_OBJETIVO, BANDERA_HORIZONTAL)
            if ev_rollo["etq_eje"] <= 0:
                if z_fallback_330 is None:
                    z_fallback_330 = z
                    motivo_imposibilidad_330 = f"Ancho {int(ancho_etq)}mm no cabe incluso en rollo estandar {int(ROLLO_BASE_MM)}mm"
                continue
            
            etq_repeat = ev_z.n * ev_rollo["etq_eje"]
            metraje = calcular_metraje_material(unidades, etq_repeat, ev_z.desarrollo_mm,
                                               ROLLO_BASE_MM, AJUSTE_DESARROLLO_REP_MM)
            
            if mejor is None or metraje["ml"] < mejor["ml"]:
                mejor = {
                    "z": z,
                    "n": ev_z.n,
                    "desarrollo_mm": ev_z.desarrollo_mm,
                    "gap_vertical": ev_z.gap,
                    "etq_eje": ev_rollo["etq_eje"],
                    "gap_horizontal_real": ev_rollo["gap_entre"],
                    "etq_repeat": etq_repeat,
                    "repeticiones": metraje["repeticiones_reales"],
                    "ml": metraje["ml"],
                    "m2": metraje["m2"],
                    "ancho_rollo_mm": ROLLO_BASE_MM,
                    "es_valido": True,
                    "rollo_alternativo": True
                }
        
        if mejor is not None:
            return mejor
        
        # Si tampoco con rollo base funcionó, retornar fallback con motivo
        if z_fallback_330 is not None:
                return {
                "z": z_fallback_330,
                "ancho_rollo_mm": ROLLO_BASE_MM,
                "es_valido": False,
                "motivo_imposibilidad": motivo_imposibilidad_330,
                "gap_horizontal_real": 0  # GAP de 0 para marcar imposibilidad
            }
    
    # Si llegamos aquí, es imposible con cualquier rollo
    if z_fallback is not None and motivo_imposibilidad:
        return {
            "z": z_fallback,
            "ancho_rollo_mm": ancho_rollo_mm,
            "es_valido": False,
            "motivo_imposibilidad": motivo_imposibilidad,
            "gap_horizontal_real": 0  # GAP de 0 para marcar imposibilidad
        }
    
    return None

# =====================================================================
# FUNCIÓN: CREAR TABLA RESUMEN VITAL
# =====================================================================
def crear_tabla_resumen(df_pedidos_componentes_stock, df_pedidos_pendientes=None, df_listademateriales=None):
    """
    Crea la tabla RESUMEN VITAL a partir de df_pedidos_componentes_stock
    
    COLUMNAS DEL RESUMEN (31 columnas - incluyendo desglose de Z):
    1. Codigo - Código del producto padre
    2. Producto_Padre - Código/nombre del producto padre
    3. Componente - Código del componente/laminado
    4. Nombre_Articulo - Nombre del artículo padre
    5. Nombre_Componente - Nombre del componente (laminado)
    6. Fecha_Entrega - Fecha de entrega del pedido
    7. Alto_mm - Alto de la etiqueta en mm
    8. Ancho_mm - Ancho de la etiqueta en mm
    9. Z_Cilindro - Cilindro óptimo calculado (Z)
    10. Desarrollo_mm - Desarrollo del cilindro en mm
    11. GAP_Desarrollo - Gap vertical en el desarrollo
    12. Etiquetas_Desarrollo - Etiquetas que caben verticalmente en el cilindro
    13. GAP_Eje - Gap horizontal en el eje
    14. Etiquetas_Eje - Etiquetas que caben horizontalmente en el eje
    15. Etiquetas_Totales - Etiquetas_Desarrollo × Etiquetas_Eje
    16. Rollo_mm - Ancho del rollo en mm
    17. Repeticiones - Número de repeticiones
    18. Metros_Lineales - ML sin merma calculados
    19. Metros_Cuadrados - M2 sin merma calculados
    20. Stock_Acumulado - Stock disponible acumulado en m2
    21. Stock_Final_m2 - Stock final (Stock - M2_sin_merma)
    22. Nota_Ventas - Cantidad pendiente de venta (desde pedidos pendientes)
    23. Factor_unidades - Cantidad del componente (desde lista de materiales) [RENOMBRADO: Cantidad → Factor_unidades]
    24. Metros_Cuadrados_Factor - Pendiente × Factor_unificado (M2 factor)
    25. Stock_Final_Factor - Stock_Acumulado - Metros_Cuadrados_Factor
    26. Indicador_Factor - 1 si Stock_Final_Factor > 0, else 0 (booleano)
    27. Disponibilidad - 1 si Stock_Final_m2 > 0, else 0
    28. Disponibilidad_Factor - 1 si Stock_Final_Factor > 0, else 0
    29-31. (Columnas adicionales si existen)
    
    Args:
        df_pedidos_componentes_stock: DataFrame maestro con datos de pedidos y componentes
        df_pedidos_pendientes: DataFrame con datos de pedidos pendientes (incluye Pendiente)
        df_listademateriales: DataFrame con lista de materiales (incluye Cantidad)
    
    Returns:
        pd.DataFrame: Tabla resumen con 31+ columnas
    """
    # Seleccionar columnas base - incluir gap si existen
    # NOTA: Agregamos 'Pedido' para usar como Numero_nota_venta
    columnas_base = [
        'Pedido',  # Se renombrará a Numero_nota_venta
        'Codigo', 'Prod_padre', 'Componente',
        'Nombre_Articulo', 'Nomb_componente',
        'FechaEntrega',
        'Etiqueta_mm_Alto', 'Etiqueta_mm_Ancho',
        'Z_Sugerido', 'Desarrollo_mm',
        'n_vertical', 'etq_eje_horizontal',
        'Ancho_rollo_mm',
        'ML_sin_merma', 'M2_sin_merma',
        'StockAcumulado'
        # NOTA: StockFinal_m2 se calcula en el código, no se importa de la BD
    ]
    
    # Agregar columnas de GAP si existen
    if 'gap_vertical' in df_pedidos_componentes_stock.columns:
        columnas_base.insert(11, 'gap_vertical')
    # Solo agregar el GAP real (horizontal) como GAP_Eje
    if 'gap_horizontal_real' in df_pedidos_componentes_stock.columns:
        # insertar en la posición correspondiente al eje
        columnas_base.insert(13, 'gap_horizontal_real')
    
    # Seleccionar solo columnas que existan en el dataframe
    columnas_disponibles = [col for col in columnas_base if col in df_pedidos_componentes_stock.columns]
    df_resumen = df_pedidos_componentes_stock[columnas_disponibles].copy()
    
    # Agregar Pendiente (Nota_Ventas) desde df_pedidos_pendientes
    if df_pedidos_pendientes is not None:
        # Merge para obtener Pendiente
        df_nota_ventas = df_pedidos_pendientes[['Codigo', 'Pendiente']].drop_duplicates()
        df_resumen = df_resumen.merge(df_nota_ventas, left_on='Codigo', right_on='Codigo', how='left')
    else:
        df_resumen['Pendiente'] = 0.0
    
    # Agregar Cantidad y Unidad_medida desde df_listademateriales
    if df_listademateriales is not None:
        # Merge para obtener Factor_unidades (renombrado desde Cantidad) y Unidad_medida
        df_factor = df_listademateriales[['Prod_padre', 'Componente', 'Cantidad', 'Unidad_medida']].drop_duplicates()
        df_resumen = df_resumen.merge(df_factor, left_on=['Prod_padre', 'Componente'], 
                                      right_on=['Prod_padre', 'Componente'], how='left')
    else:
        df_resumen['Cantidad'] = 0.0
        df_resumen['Unidad_medida'] = ''
    
    # Convertir a numéricos y llenar NaN con 0
    df_resumen['Pendiente'] = pd.to_numeric(df_resumen['Pendiente'], errors='coerce').fillna(0)
    df_resumen['Cantidad'] = pd.to_numeric(df_resumen['Cantidad'], errors='coerce').fillna(0)
    df_resumen['M2_sin_merma'] = pd.to_numeric(df_resumen['M2_sin_merma'], errors='coerce').fillna(0)
    df_resumen['StockAcumulado'] = pd.to_numeric(df_resumen['StockAcumulado'], errors='coerce').fillna(0)
    df_resumen['FechaEntrega'] = pd.to_datetime(df_resumen['FechaEntrega'], errors='coerce')
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # NUEVA COLUMNA: ETIQUETAS_A_PRODUCIR
    # ═══════════════════════════════════════════════════════════════════════════════
    # Si Unidad_Medida = UNIDAD → Pendiente
    # Si Unidad_Medida = MILES o MIL → Pendiente × 1000
    df_resumen['Etiquetas_a_Producir'] = df_resumen.apply(
        lambda row: (row['Pendiente'] * 1000 
                    if str(row.get('Unidad_medida', '')).strip().upper() in ['MILES', 'MIL']
                    else row['Pendiente']),
        axis=1
    )
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # VALIDACION CRITICA: Limpiar Stock_Acumulado negativo en primeras apariciones
    # ═══════════════════════════════════════════════════════════════════════════════
    # Si el STOCK INICIAL de un componente es negativo, convertir a 0
    # Esto indica que el componente NO existe en BD pero tiene valor errado asignado
    negativos_iniciales = 0
    for componente_unico in df_resumen['Componente'].unique():
        filas_componente = df_resumen[df_resumen['Componente'] == componente_unico]
        if len(filas_componente) > 0:
            idx_primera = filas_componente.index[0]
            stock_inicial = df_resumen.at[idx_primera, 'StockAcumulado']
            if stock_inicial < 0:
                print(f"[ADVERTENCIA] Stock inicial negativo para componente {componente_unico}: {stock_inicial:.4f} → Reemplazando con 0", flush=True)
                df_resumen.at[idx_primera, 'StockAcumulado'] = 0
                negativos_iniciales += 1
    
    if negativos_iniciales > 0:
        print(f"[INFO] Se corrigieron {negativos_iniciales} componentes con stock inicial NEGATIVO", flush=True)
    
    # CASCADEO DE STOCK POR COMPONENTE Y FECHA
    # Ordenar por Componente y Fecha_Entrega (más antigua primero)
    df_resumen = df_resumen.sort_values(by=['Componente', 'FechaEntrega']).reset_index(drop=True)
    
    # Calcular Metros_Cuadrados_Con_Merma = Metros_Cuadrados × 1.12 (12% extra de merma)
    # Este cálculo se hace ANTES del cascadeo porque se necesita para la fórmula tradicional
    if 'M2_sin_merma' in df_resumen.columns:
        df_resumen['Metros_Cuadrados_Con_Merma'] = df_resumen['M2_sin_merma'] * 1.12
    else:
        df_resumen['Metros_Cuadrados_Con_Merma'] = 0.0
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # CALCULAR FACTOR_UNIFICADO (ANTES de Metros_Cuadrados_Factor)
    # ═══════════════════════════════════════════════════════════════════════════════
    # NUEVA REGLA DE NORMALIZACIÓN:
    # 1. Si Unidad_Medida = 'UNIDAD': mantener Cantidad sin cambios (EXCEPTO si Cantidad > 1)
    # 2. Si Cantidad > 1 (independiente de la unidad): dividir entre 1000
    # 3. Si no: mantener Cantidad como está
    
    # DEBUG: Verificar valores de Cantidad y Unidad_medida ANTES de calcular
    print("\n[DEBUG] VERIFICACION - Factor_unificado (NUEVA REGLA)", flush=True)
    print("─" * 100, flush=True)
    print(f"Cantidad (min): {df_resumen['Cantidad'].min()} | (max): {df_resumen['Cantidad'].max()}", flush=True)
    print(f"Unidad_medida UNIQUE: {df_resumen['Unidad_medida'].unique()}", flush=True)
    print("\nPrimeros 10 registros:", flush=True)
    print(df_resumen[['Cantidad', 'Unidad_medida']].head(10).to_string(), flush=True)
    print("─" * 100, flush=True)
    
    df_resumen['Factor_unificado'] = df_resumen.apply(
        lambda row: (row['Cantidad'] / 1000 if row['Cantidad'] > 1 else row['Cantidad'])
                    if str(row.get('Unidad_medida', '')).strip().upper() == 'UNIDAD'
                    else (row['Cantidad'] / 1000 if row['Cantidad'] > 1 else row['Cantidad']),
        axis=1
    )
    
    # Mostrar resultado del Factor_unificado
    print("\n[DEBUG] RESULTADO - Factor_unificado calculado (NUEVA REGLA)", flush=True)
    print("─" * 100, flush=True)
    print(df_resumen[['Cantidad', 'Unidad_medida', 'Factor_unificado']].head(10).to_string(), flush=True)
    print("─" * 100, flush=True)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # NUEVA COLUMNA: METROS_CUADRADOS_FACTOR_SISTEMA
    # ═══════════════════════════════════════════════════════════════════════════════
    # Metros_Cuadrados_Factor_Sistema = Pendiente × Cantidad (directo, sin normalizar)
    df_resumen['Metros_Cuadrados_Factor_Sistema'] = df_resumen['Pendiente'] * df_resumen['Cantidad']
    
    # Calcular Metros_Cuadrados_Factor = Pendiente × Factor_unificado (DESPUÉS de Factor_unificado)
    df_resumen['Metros_Cuadrados_Factor'] = df_resumen['Pendiente'] * df_resumen['Factor_unificado']
    
    # Calcular Etiquetas_Totales = n_vertical × etq_eje_horizontal
    if 'n_vertical' in df_resumen.columns and 'etq_eje_horizontal' in df_resumen.columns:
        df_resumen['Etiquetas_Totales'] = (df_resumen['n_vertical'] * df_resumen['etq_eje_horizontal'])
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # NORMALIZAR STOCK SEGÚN UNIDAD_MEDIDA
    # ═══════════════════════════════════════════════════════════════════════════════
    # El stock de BD viene en su unidad nativa, debemos normalizarlo a M²:
    # - Si UNIDAD: Stock × 1000 (cada unidad = 1000 m²)
    # - Si ROLLO: Stock × Factor_de_conversion (m² por rollo)
    # - Si MILES: Stock × Factor_unificado (factor normalizado)
    
    print("\n[DEBUG] NORMALIZACIÓN DE STOCK POR UNIDAD_MEDIDA", flush=True)
    print("─" * 100, flush=True)
    
    def normalizar_stock(row):
        """
        Normaliza el stock según la Unidad_Medida del componente:
        - UNIDAD/UN/MILES/MIL/vacío → Stock × 1000 (convertir a base)
        - ROLLO/ROLLOS → Stock × Factor_unificado (usar factor de conversión)
        """
        stock_original = row['StockAcumulado']
        unidad = str(row.get('Unidad_medida', '')).strip().upper()
        
        # Si es UNIDAD/UN, MILES/MIL o está vacío → multiplicar por 1000
        if unidad in ['UNIDAD', 'UN', 'UNI', 'MILES', 'MIL', 'MILE', '']:
            return stock_original * 1000
        
        # Si es ROLLO/ROLLOS → usar Factor_unificado (m² por rollo)
        elif unidad in ['ROLLO', 'ROLLOS', 'ROL', 'ROLOS']:
            factor_rollo = row.get('Factor_unificado', 1)
            return stock_original * factor_rollo
        
        # Por defecto → asumir MILES y multiplicar por 1000
        else:
            return stock_original * 1000
    
    df_resumen['StockAcumulado'] = df_resumen.apply(normalizar_stock, axis=1)
    
    print("  ✓ Stock normalizado según Unidad_Medida", flush=True)
    print(df_resumen[['Componente', 'Unidad_medida', 'Factor_unificado', 'StockAcumulado']].drop_duplicates(subset=['Componente']).head(10).to_string(), flush=True)
    print("─" * 100, flush=True)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # CASCADEO DUAL DE STOCK - DOS LÓGICAS PARALELAS
    # ═══════════════════════════════════════════════════════════════════════════════
    
    # Guardar ANTES de modificar: Stock_Acumulado original (de BD, ya normalizado)
    df_resumen['Stock_Acumulado_Original'] = df_resumen['StockAcumulado'].copy()
    
    # Inicializar columnas para cascadeo
    df_resumen['Stock_Acumulado_Cascada'] = 0.0        # Para fórmula tradicional (izquierda)
    df_resumen['Stock_Acumulado_Factor_Cascada'] = 0.0 # Para factor (derecha)
    df_resumen['Stock_Final_m2'] = 0.0                 # Resultado lado izquierdo
    df_resumen['Stock_Final_Factor'] = 0.0             # Resultado lado derecho
    
    componente_anterior = None
    stock_tradicional = 0.0      # Para lado izquierdo (Motor)
    stock_factor = 0.0           # Para lado derecho (Factor)
    
    for idx in df_resumen.index:
        componente_actual = df_resumen.at[idx, 'Componente']
        
        # Si cambió el componente, reiniciar con los Stock_Acumulado originales (de BD)
        if componente_actual != componente_anterior:
            stock_original = df_resumen.at[idx, 'Stock_Acumulado_Original']
            stock_tradicional = stock_original
            stock_factor = stock_original
            componente_anterior = componente_actual
        
        # ───────────────────────────────────────────────────────────────────────────
        # LADO IZQUIERDO: FÓRMULA TRADICIONAL (Motor - Metros con Merma)
        # ───────────────────────────────────────────────────────────────────────────
        df_resumen.at[idx, 'Stock_Acumulado_Cascada'] = stock_tradicional
        metros_merma = df_resumen.at[idx, 'Metros_Cuadrados_Con_Merma']
        stock_final_m2 = stock_tradicional - metros_merma
        df_resumen.at[idx, 'Stock_Final_m2'] = stock_final_m2
        
        # Actualizar stock para la siguiente fila (usando stock final m2)
        stock_tradicional = stock_final_m2
        
        # ───────────────────────────────────────────────────────────────────────────
        # LADO DERECHO: CÁLCULOS POR FACTOR (Metros Factor)
        # ───────────────────────────────────────────────────────────────────────────
        df_resumen.at[idx, 'Stock_Acumulado_Factor_Cascada'] = stock_factor
        metros_factor = df_resumen.at[idx, 'Metros_Cuadrados_Factor']
        stock_final_factor = stock_factor - metros_factor
        df_resumen.at[idx, 'Stock_Final_Factor'] = stock_final_factor
        
        # Actualizar stock para la siguiente fila (usando stock final factor)
        stock_factor = stock_final_factor
    
    # Asignar valores cascadeados al Stock_Acumulado (para lado izquierdo)
    df_resumen['StockAcumulado'] = df_resumen['Stock_Acumulado_Cascada']
    df_resumen = df_resumen.drop(columns=['Stock_Acumulado_Cascada'])
    
    # Crear Stock_Acumulado_Factor (para lado derecho)
    df_resumen['Stock_Acumulado_Factor'] = df_resumen['Stock_Acumulado_Factor_Cascada']
    df_resumen = df_resumen.drop(columns=['Stock_Acumulado_Factor_Cascada', 'Stock_Acumulado_Original'])
    
    # Calcular Indicador_Factor: 1 si Stock_Final_Factor > 0, else 0 (booleano)
    df_resumen['Indicador_Factor'] = (df_resumen['Stock_Final_Factor'] > 0).astype(bool)
    
    # Renombrar columnas para mayor claridad
    renombrados = {
        'Pedido': 'Numero_nota_venta',
        'Codigo': 'Cod_lista_de_productos',
        'Prod_padre': 'Producto_Padre',
        'Componente': 'Componente',
        'Nombre_Articulo': 'Nombre_Articulo',
        'Nomb_componente': 'Nombre_Componente',
        'FechaEntrega': 'Fecha_Entrega',
        'Etiqueta_mm_Alto': 'Alto_mm',
        'Etiqueta_mm_Ancho': 'Ancho_mm',
        'Z_Sugerido': 'Z_Cilindro',
        'Desarrollo_mm': 'Desarrollo_mm',
        'gap_vertical': 'GAP_Desarrollo',
        'n_vertical': 'Etiquetas_Desarrollo',
    'gap_horizontal_real': 'GAP_Eje',
        'etq_eje_horizontal': 'Etiquetas_Eje',
        'Ancho_rollo_mm': 'Rollo_mm',
        'ML_sin_merma': 'Metros_Lineales',
        'M2_sin_merma': 'Metros_Cuadrados',
        'StockAcumulado': 'Stock_Acumulado',
        'StockFinal_m2': 'Stock_Final_m2',
        'Cantidad': 'Factor_unidades',
        'Unidad_medida': 'Unidad_Medida'
    }
    
    # Solo renombrar columnas que existen
    renombrados_existentes = {k: v for k, v in renombrados.items() if k in df_resumen.columns}
    df_resumen = df_resumen.rename(columns=renombrados_existentes)
    
    # Agregar columnas de disponibilidad (basadas en stock final)
    # Disponibilidad: basada en Stock_Final_m2 (lado izquierdo - Motor)
    df_resumen['Disponibilidad'] = (df_resumen['Stock_Final_m2'] > 0).astype(int)
    # Disponibilidad_Factor: basada en Stock_Final_Factor (lado derecho - Factor)
    df_resumen['Disponibilidad_Factor'] = (df_resumen['Stock_Final_Factor'] > 0).astype(int)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # REORDENAR COLUMNAS POR BLOQUES LÓGICOS (Diferenciación clara)
    # ═══════════════════════════════════════════════════════════════════════════════
    
    orden_deseado = [
        # ═══════════════════════════════════════════════════════════════════════════════
        # LADO IZQUIERDO: MOTOR DE CÁLCULO (Basado en dimensiones físicas)
        # ═══════════════════════════════════════════════════════════════════════════════
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 0: DISPONIBILIDAD - LO PRIMERO QUE VE EL USUARIO
        # ───────────────────────────────────────────────────────────────────────────
        'Disponibilidad',           # ¿Hay stock final después de merma? (1=Sí, 0=No)
        'Disponibilidad_Factor',    # ¿Hay stock final? (1=Sí, 0=No) - LADO FACTOR
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 1: IDENTIFICACIÓN - ¿QUÉ ES ESTE REGISTRO?
        # ───────────────────────────────────────────────────────────────────────────
        'Numero_nota_venta',        # Referencia del pedido (DUPLICADO en ambos lados)
        'Codigo',                   # Código producto padre
        'Producto_Padre',           # Nombre producto padre
        'Componente',               # Código del componente (laminado)
        'Nombre_Articulo',          # Nombre artículo
        'Nombre_Componente',        # Nombre componente (laminado)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 2: CONTEXTO DEL PEDIDO - ¿QUÉ SE NECESITA?
        # ───────────────────────────────────────────────────────────────────────────
        'Fecha_Entrega',            # Cuándo se necesita (DUPLICADO en ambos lados)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 3: DIMENSIONES DE ETIQUETA - ¿QUÉ TAMAÑO?
        # ───────────────────────────────────────────────────────────────────────────
        'Alto_mm',                  # Alto de etiqueta en mm
        'Ancho_mm',                 # Ancho de etiqueta en mm
        'Rollo_mm',                 # Ancho del rollo disponible
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 4: MOTOR DE CÁLCULO - EJE VERTICAL (Cilindro)
        # ───────────────────────────────────────────────────────────────────────────
        'Z_Cilindro',               # Cilindro elegido (60-168)
        'Desarrollo_mm',            # Desarrollo del cilindro (Z × 3.175 mm)
        'Etiquetas_Desarrollo',     # Cuántas etiquetas caben verticalmente
        'GAP_Desarrollo',           # GAP vertical entre etiquetas
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 5: MOTOR DE CÁLCULO - EJE HORIZONTAL (Rollo)
        # ───────────────────────────────────────────────────────────────────────────
        'Etiquetas_Eje',            # Etiquetas que caben horizontalmente
        'GAP_Eje',                  # GAP horizontal entre etiquetas
        'Etiquetas_Totales',        # Total (Etiquetas_Desarrollo × Etiquetas_Eje)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 5A: ETIQUETAS A PRODUCIR - NUEVA COLUMNA
        # ───────────────────────────────────────────────────────────────────────────
        'Etiquetas_a_Producir',     # Pendiente (si UNIDAD) o Pendiente×1000 (si MILES/MIL)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 6: METRAJE DEL MOTOR (Resultados del motor sin merma)
        # ───────────────────────────────────────────────────────────────────────────
        'Metros_Lineales',          # ML sin merma necesarios
        'Metros_Cuadrados',         # M2 sin merma necesarios
        'Metros_Cuadrados_Con_Merma', # M2 con merma (Metros_Cuadrados × 1.12)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 7: STOCK - MOTOR DE CÁLCULO
        # ───────────────────────────────────────────────────────────────────────────
        'Stock_Acumulado',          # Stock disponible al iniciar (cascadeado)
        'Stock_Final_m2',           # Stock final (Stock_Acumulado - Metros_Cuadrados_Con_Merma)
        
        # ═══════════════════════════════════════════════════════════════════════════════
        # LADO DERECHO: CÁLCULOS POR FACTOR (Basado en Pendiente × Factor_unificado)
        # ═══════════════════════════════════════════════════════════════════════════════
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 8: IDENTIFICACIÓN (Duplicado para contexto de Factor)
        # ───────────────────────────────────────────────────────────────────────────
        'Numero_nota_venta',        # Referencia del pedido (DUPLICADO)
        'Fecha_Entrega',            # Cuándo se necesita (DUPLICADO)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 9: CANTIDADES - COMPONENTES CLAVE DEL FACTOR
        # ───────────────────────────────────────────────────────────────────────────
        'Pendiente',                # Cantidad pendiente de venta (CLAVE)
        'Factor_unidades',          # Cantidad componente por unidad (CLAVE) [RENOMBRADO]
        'Unidad_Medida',            # Unidad de medida del componente (NUEVO)
        'Factor_unificado',         # Factor normalizado: Si Cant>1 entonces ÷1000, si no mantener igual (NUEVO)
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 10: CONSUMO POR FACTOR (Lógica de venta)
        # ───────────────────────────────────────────────────────────────────────────
        'Metros_Cuadrados_Factor',  # M2 que se consumirá (Pendiente × Factor_unificado)
        'Metros_Cuadrados_Factor_Sistema',  # M2 sin normalización (Pendiente × Cantidad) - NUEVA COLUMNA
        
        # ───────────────────────────────────────────────────────────────────────────
        # BLOQUE 11: STOCK FINAL - FACTOR
        # ───────────────────────────────────────────────────────────────────────────
        'Stock_Acumulado_Factor',   # Stock disponible al iniciar (cascadeado) (DUPLICADO)
        'Stock_Final_Factor',       # Stock final (Stock_Acumulado - Metros_Cuadrados_Factor)
        'Indicador_Factor'          # Booleano: ¿hay stock final?
    ]
    
    # Permitir columnas duplicadas para mostrar ambos lados (Motor + Factor)
    # Validar que todas las columnas en orden_deseado existan, filtrar solo las que existen
    columnas_existentes = set(df_resumen.columns)
    orden_final = [col for col in orden_deseado if col in columnas_existentes]
    
    # Agregar cualquier columna restante que no esté en orden_final
    columnas_restantes = [col for col in df_resumen.columns if col not in columnas_existentes or 
                          col not in set(orden_deseado)]
    
    # Crear nueva tabla con columnas en el orden deseado (permitiendo duplicados)
    df_resumen_ordenado = pd.DataFrame()
    for col in orden_final:
        df_resumen_ordenado[col] = df_resumen[col]
    
    # Agregar columnas restantes
    for col in columnas_restantes:
        if col not in df_resumen_ordenado.columns:
            df_resumen_ordenado[col] = df_resumen[col]
    
    df_resumen = df_resumen_ordenado
    
    return df_resumen

# Configuración de conexión a SQL Server
server = '10.101.2.181'
database = 'SAP_G02E05_Innoprint'
username = 'ReportesInnoprint'
password = 'm^9S*^N$v2AR'

# Conexión a SQL Server usando SQLAlchemy CON OPTIMIZACIONES
# ═══════════════════════════════════════════════════════════════════════════════
conn_str = (
    f"mssql+pyodbc://{username}:{password}@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server"
)

try:
    print("[INFO] Creando conexión optimizada a SQL Server...", flush=True)
    
    # OPTIMIZACIONES:
    # - pool_size: número de conexiones reutilizables (default 5)
    # - max_overflow: conexiones adicionales si se agotan (default 10)
    # - pool_recycle: recicla conexiones cada 3600 segundos (evita timeouts)
    # - echo: False para reducir verbosidad
    from sqlalchemy.pool import QueuePool
    
    engine = create_engine(
        conn_str,
        poolclass=QueuePool,
        pool_size=5,
        max_overflow=10,
        pool_recycle=3600,  # Reciclar conexiones cada hora
        connect_args={
            'timeout': 30,  # Timeout de conexión: 30 segundos
            'fast_executemany': True,  # Ejecutar múltiples comandos rápidamente
        },
        echo=False
    )
    
    print("[INFO] Engine creado ✓", flush=True)
    print("[INFO] Conectando a SQL Server y ejecutando SELECT FROM vista_stock_General...", flush=True)
    
    # OPTIMIZACIÓN: Query mejorada con índices y filtros más específicos
    query = """
    SELECT TOP 50000 * 
    FROM dbo.vista_stock_General AS stock 
    WHERE Categoria LIKE '%LAMINADO%' 
       OR ItemName LIKE '%LAMINADO%'
    """
    
    df_stock = pd.read_sql(query, engine)
    print(f"[OK] Datos de stock cargados ✓ ({len(df_stock)} registros)", flush=True)

    print("[INFO] Cargando lista de materiales...", flush=True)
    query_listademateriales = """
    SELECT TOP 100 PERCENT 
        T0.Code AS Prod_padre, 
        T2.ItemName, 
        T0.CreateDate AS Fec_creacion, 
        T0.UpdateDate AS Fec_acrtual, 
        T0.ToWH AS Bodega_principal, 
        T5.WhsName AS Nomb_bod_prin, 
        T1.Code AS Componente,
        T2.InvntryUom AS Unidad_medida,
        T3.ItemName AS Nomb_componente, 
        T1.Quantity AS Cantidad,
        T1.Warehouse AS Bod_componente, 
        T6.WhsName AS Nom_Bod_comp, 
        T4.U_NAME AS Usuario_creac
    FROM dbo.OITT T0
        INNER JOIN dbo.ITT1 T1 ON T0.Code = T1.Father
        INNER JOIN dbo.OITM T2 ON T1.Father = T2.ItemCode
        INNER JOIN dbo.OITM T3 ON T1.Code = T3.ItemCode
        INNER JOIN dbo.OUSR T4 ON T0.UserSign = T4.INTERNAL_K
        INNER JOIN dbo.OWHS T5 ON T0.ToWH = T5.WhsCode
        INNER JOIN dbo.OWHS T6 ON T1.Warehouse = T6.WhsCode
    ORDER BY Cantidad DESC
    """
    print("[DEBUG] Ejecutando query de lista de materiales...", flush=True)
    df_listademateriales = pd.read_sql(query_listademateriales, engine)
    print(f"[OK] Lista de materiales cargada ({len(df_listademateriales)} registros) ✓", flush=True)
    df_listademateriales['Primer_palabra_componente'] = df_listademateriales['Nomb_componente'].str.split(' ').str[0]
    df_listademateriales['Union_palabra_padre'] = df_listademateriales['Prod_padre'].astype(str) + df_listademateriales['Primer_palabra_componente']
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # AGREGAR COLUMNA STOCK_FINAL A DF_STOCK
    # ═══════════════════════════════════════════════════════════════════════════════
    # Merge df_stock con df_listademateriales para traer Unidad_medida
    df_stock_con_unidad = pd.merge(
        df_stock,
        df_listademateriales[['Componente', 'Unidad_medida']].drop_duplicates(),
        left_on='ItemCode',
        right_on='Componente',
        how='left'
    )
    
    # Función para calcular Stock_Final según Unidad_Medida
    def calcular_stock_final(row):
        """
        Calcula Stock_Final respetando la Unidad_Medida:
        - Si UNIDAD/UN: Stock × 1000 (convertir a base)
        - Si MILES/MIL: Stock × 1000 (MILES implica que está en miles, hay que expandir)
        - Si ROLLO/ROLLOS: mantener igual (el stock está en rollos, unidad especial)
        - Si está vacío: asumir MILES y multiplicar × 1000
        """
        cantidad = row.get('Cantidad', 0)
        unidad = str(row.get('Unidad_medida', '')).strip().upper()
        
        # Si es UNIDAD/UN, MILES/MIL o está vacío → multiplicar por 1000
        if unidad in ['UNIDAD', 'UN', 'UNI', 'MILES', 'MIL', 'MILE', '']:
            # Todos estos casos se multiplican por 1000 para convertir a la unidad base
            return cantidad * 1000
        
        # Normalizar variaciones de "ROLLO" → mantener igual
        elif unidad in ['ROLLO', 'ROLLOS', 'ROL', 'ROLOS']:
            # Si es ROLLO/ROLLOS, mantener igual (el stock está en rollos)
            return cantidad
        
        else:
            # Por defecto (cualquier otra cosa), asumir MILES y multiplicar por 1000
            return cantidad * 1000
    
    print("\n[DEBUG] CALCULANDO STOCK_FINAL EN DF_STOCK", flush=True)
    print("─" * 100, flush=True)
    
    df_stock_con_unidad['Stock_Final'] = df_stock_con_unidad.apply(calcular_stock_final, axis=1)
    
    print("  ✓ Stock_Final calculado según Unidad_Medida", flush=True)
    print(df_stock_con_unidad[['ItemCode', 'Unidad_medida', 'Cantidad', 'Stock_Final']].drop_duplicates(subset=['ItemCode']).head(10).to_string(), flush=True)
    print("─" * 100, flush=True)
    
    # Actualizar df_stock con la nueva columna (usar Stock_Final en lugar de Cantidad)
    df_stock = df_stock_con_unidad.copy()

    print("[INFO] Cargando pedidos pendientes...", flush=True)
    try:
        query_pedidos_pendientes = "SELECT * FROM SAP_G02E05_Innoprint.dbo.Vista_pedidos_pendientes"
        # Intentar cargar con timeout de 120 segundos (más generoso)
        df_pedidos_pendientes = pd.read_sql(query_pedidos_pendientes, engine, timeout=120)
        print("[OK] Pedidos pendientes cargados", flush=True)
    except Exception as e:
        print(f"[ADVERTENCIA] No se pudieron cargar pedidos pendientes: {e}", flush=True)
        print("[INFO] Continuando con DataFrame vacío...", flush=True)
        df_pedidos_pendientes = pd.DataFrame()  # DataFrame vacío para evitar que se cuelgue
    # --- EXTRAER MEDIDAS DE ETIQUETA ---
    import re
    
    # Función mejorada para extraer medidas que maneja puntos iniciales errados
    def extraer_medidas_stock(articulo):
        """
        Extrae medidas de etiqueta del nombre del artículo.
        Maneja casos como ".219.5" ignorando puntos iniciales errados.
        
        Busca patrones:
        - Dos medidas: 55x66, 55,5 x 66,7, .219.5x100 (ignora primer punto)
        - Una medida: 170 MM, 170mm, .170.5 MM, etc.
        
        Returns:
            Series([alto, ancho]) o Series([0, 0]) si no encuentra
        """
        if not isinstance(articulo, str):
            return pd.Series([0, 0])
        
        # Limpiar puntos iniciales errados: ".219.5" → "219.5"
        # Caso 1: Punto inicial al comienzo
        articulo_limpio = re.sub(r'^\.(\d)', r'\1', articulo)
        # Caso 2: Punto inicial después de un espacio
        articulo_limpio = re.sub(r'(\s)\.(\d)', r'\1\2', articulo_limpio)
        
        # Primero busca el patrón con dos medidas: 55x66, 55,5 x 66,7, 55.5 X 66.7, .219.5x100, etc.
        match_dos_medidas = re.search(r'([\d,\.]+)\s*[xX]\s*([\d,\.]+)', articulo_limpio)
        if match_dos_medidas:
            try:
                # Convertir comas a puntos
                alto_str = match_dos_medidas.group(1).replace(',', '.')
                ancho_str = match_dos_medidas.group(2).replace(',', '.')
                
                # Limpiar puntos múltiples al inicio: ".219.5" → "219.5"
                alto_str = re.sub(r'^\.+', '', alto_str)
                ancho_str = re.sub(r'^\.+', '', ancho_str)
                
                alto = float(alto_str)
                ancho = float(ancho_str)
                return pd.Series([alto, ancho])
            except ValueError:
                pass  # Si falla, intentar con una medida
        
        # Si no encuentra dos medidas, busca una sola medida acompañada de MM
        # Patrones: "170 MM", "170MM", "170 mm", "170mm", ".170.5 MM", etc.
        match_una_medida = re.search(r'([\d,\.]+)\s*(?:mm|MM|Mm|mM)\b', articulo_limpio, re.IGNORECASE)
        if match_una_medida:
            try:
                medida_str = match_una_medida.group(1).replace(',', '.')
                # Limpiar puntos iniciales
                medida_str = re.sub(r'^\.+', '', medida_str)
                medida = float(medida_str)
                # Si encuentra una sola medida con MM, asume que alto y ancho son iguales
                return pd.Series([medida, medida])
            except ValueError:
                pass
        
        return pd.Series([0, 0])

    # Crear la tabla stock acumulado
    df_stock_acumulado = (
        df_stock.groupby(['ItemCode', 'ItemName', 'Categoria'], as_index=False)['Stock_Final']
        .sum()
        .rename(columns={'Stock_Final': 'StockAcumulado'})
    )
    df_stock_acumulado['StockAcumulado'] = df_stock_acumulado['StockAcumulado'].astype(int)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # VALIDACION: Si Stock_Acumulado es negativo, reemplazar con 0
    # ═══════════════════════════════════════════════════════════════════════════════
    negativos_antes = (df_stock_acumulado['StockAcumulado'] < 0).sum()
    if negativos_antes > 0:
        print(f"\n[ADVERTENCIA] Se encontraron {negativos_antes} componentes con StockAcumulado NEGATIVO", flush=True)
        print("[INFO] Reemplazando valores negativos con 0", flush=True)
        df_stock_acumulado['StockAcumulado'] = df_stock_acumulado['StockAcumulado'].clip(lower=0)
    
    # Filtrar solo los que comienzan con 'LAMINADO' en ItemName, ignorando nulos
    df_stock_acumulado = df_stock_acumulado[
        df_stock_acumulado['ItemName'].fillna('').str.startswith('LAMINADO')
    ]

    # Agregar columna con todas las fechas de compra (FechAdmis) concatenadas por ItemCode
    fechas_por_item = (
        df_stock[df_stock['ItemName'].fillna('').str.startswith('LAMINADO')]
        .groupby('ItemCode')['FechAdmis']
        .apply(lambda fechas: ' - '.join(pd.to_datetime(fechas.dropna()).dt.strftime('%d-%m-%Y')))
        .reset_index()
        .rename(columns={'FechAdmis': 'FechasCompra'})
    )
    df_stock_acumulado = df_stock_acumulado.merge(fechas_por_item, on='ItemCode', how='left')

    # Si el stock acumulado es 0, poner la fecha actual en FechasCompra
    fecha_actual = pd.Timestamp.now().strftime('%d-%m-%Y')
    df_stock_acumulado.loc[df_stock_acumulado['StockAcumulado'] == 0, 'FechasCompra'] = fecha_actual

    # --- NUEVO: Left join entre pedidos pendientes y lista de materiales ---
    # Asumimos que la columna de código en pedidos pendientes se llama 'Codigo'
    # Si tiene otro nombre, reemplaza 'Codigo' por el nombre correcto
    df_pedidos_con_componentes = pd.merge(
        df_pedidos_pendientes,
        df_listademateriales[['Prod_padre', 'Componente', 'Nomb_componente']],
        left_on='Codigo',  # Cambia 'Codigo' si el nombre es diferente
        right_on='Prod_padre',
        how='left'
    )

    # --- Procesamiento paso a paso de df_pedidos_componentes_stock ---
    # 1. Merge con stock acumulado
    df_pedidos_componentes_stock = pd.merge(
        df_pedidos_con_componentes,
        df_stock_acumulado[['ItemCode', 'StockAcumulado']],
        left_on='Componente',
        right_on='ItemCode',
        how='left'
    )
    df_pedidos_componentes_stock = df_pedidos_componentes_stock.drop(columns=['ItemCode'])

    # 2. Filtrar solo componentes que empiezan por 'LAMINADO'
    df_pedidos_componentes_stock = df_pedidos_componentes_stock[
        df_pedidos_componentes_stock['Nomb_componente'].fillna('').str.startswith('LAMINADO')
    ]

    # 3. Limpiar nulos y convertir a float
    df_pedidos_componentes_stock['StockAcumulado'] = df_pedidos_componentes_stock['StockAcumulado'].fillna(0).astype(float)
    df_pedidos_componentes_stock['Pendiente'] = df_pedidos_componentes_stock['Pendiente'].fillna(0).astype(float)

    # 4. Extraer medidas de etiqueta
    posibles_col = ["nombre_articulo", "Nombre_articulo", "Nombre_Articulo", "articulo", "Articulo", "NOMBRE_ARTICULO"]
    col_medidas = next((col for col in posibles_col if col in df_pedidos_componentes_stock.columns), None)
    def extraer_medidas_stock(articulo):
        """
        Extrae medidas de etiqueta del nombre del artículo.
        Maneja casos como ".219.5" ignorando puntos iniciales errados.
        
        Busca patrones:
        - Dos medidas: 55x66, 55,5 x 66,7, .219.5x100 (ignora primer punto)
        - Una medida: 170 MM, 170mm, .170.5 MM, etc.
        
        Returns:
            Series([alto, ancho]) o Series([0, 0]) si no encuentra
        """
        if not isinstance(articulo, str):
            return pd.Series([0, 0])
        
        # Limpiar puntos iniciales errados: ".219.5" → "219.5"
        # Caso 1: Punto inicial al comienzo
        articulo_limpio = re.sub(r'^\.(\d)', r'\1', articulo)
        # Caso 2: Punto inicial después de un espacio
        articulo_limpio = re.sub(r'(\s)\.(\d)', r'\1\2', articulo_limpio)
        
        # Primero busca el patrón con dos medidas: 55x66, 55,5 x 66,7, 55.5 X 66.7, .219.5x100, etc.
        match_dos_medidas = re.search(r'([\d,\.]+)\s*[xX]\s*([\d,\.]+)', articulo_limpio)
        if match_dos_medidas:
            try:
                # Convertir comas a puntos
                alto_str = match_dos_medidas.group(1).replace(',', '.')
                ancho_str = match_dos_medidas.group(2).replace(',', '.')
                
                # Limpiar puntos múltiples al inicio: ".219.5" → "219.5"
                alto_str = re.sub(r'^\.+', '', alto_str)
                ancho_str = re.sub(r'^\.+', '', ancho_str)
                
                alto = float(alto_str)
                ancho = float(ancho_str)
                return pd.Series([alto, ancho])
            except ValueError:
                pass  # Si falla, intentar con una medida
        
        # Si no encuentra dos medidas, busca una sola medida acompañada de MM
        # Patrones: "170 MM", "170MM", "170 mm", "170mm", ".170.5 MM", etc.
        match_una_medida = re.search(r'([\d,\.]+)\s*(?:mm|MM|Mm|mM)\b', articulo_limpio, re.IGNORECASE)
        if match_una_medida:
            try:
                medida_str = match_una_medida.group(1).replace(',', '.')
                # Limpiar puntos iniciales
                medida_str = re.sub(r'^\.+', '', medida_str)
                medida = float(medida_str)
                # Si encuentra una sola medida con MM, asume que alto y ancho son iguales
                return pd.Series([medida, medida])
            except ValueError:
                pass
        
        return pd.Series([0, 0])
    if col_medidas:
        df_pedidos_componentes_stock[['Etiqueta_Alto', 'Etiqueta_Ancho']] = df_pedidos_componentes_stock[col_medidas].apply(extraer_medidas_stock)
    if 'Etiqueta_Alto' not in df_pedidos_componentes_stock.columns:
        df_pedidos_componentes_stock['Etiqueta_Alto'] = 0
    if 'Etiqueta_Ancho' not in df_pedidos_componentes_stock.columns:
        df_pedidos_componentes_stock['Etiqueta_Ancho'] = 0

    # 5. Calcular área en m2
    df_pedidos_componentes_stock['Etiqueta_m2'] = (
        df_pedidos_componentes_stock['Etiqueta_Alto'].astype(float) * df_pedidos_componentes_stock['Etiqueta_Ancho'].astype(float)
    ) / 1_000_000

    # 5a. CREAR ETIQUETAS_A_PRODUCIR ANTES DEL MOTOR
    # Etiquetas_a_Producir = Pendiente x 1000 si UMInv es MILES, sino Pendiente
    if 'Etiquetas_a_Producir' not in df_pedidos_componentes_stock.columns:
        df_pedidos_componentes_stock['Etiquetas_a_Producir'] = df_pedidos_componentes_stock.apply(
            lambda row: (row['Pendiente'] * 1000 
                        if str(row.get('UMInv', '')).strip().upper() in ['MILES', 'MIL', 'MILE']
                        else row['Pendiente']),
            axis=1
        )

    # 5b. MOTOR DE CALCULO: Obtener Z sugerido, metraje ML y M2
    print("Ejecutando motor de cálculo para obtener Z sugerido y metraje...", flush=True)
    
    # Lista para rastrear impossibilidades técnicas
    imposibilidades_tecnicas = []
    
    def calcular_z_y_metraje(row):
        """Calcula Z sugerido, ML y M2 para cada componente"""
        alto = row['Etiqueta_Alto']
        ancho = row['Etiqueta_Ancho']
        # CAMBIO CRITICO: Usar Etiquetas_a_Producir en lugar de Pendiente
        # Etiquetas_a_Producir se crea ANTES en la linea ~1205
        etiquetas_a_producir = row['Etiquetas_a_Producir']
        nombre_componente = row.get('Nomb_componente', '')
        codigo = row.get('Codigo', '')
        
        # Extraer ancho del rollo desde Nombre_Componente
        ancho_rollo = extraer_ancho_rollo_mm(nombre_componente)
        if ancho_rollo is None:
            ancho_rollo = ROLLO_BASE_MM  # Usar valor por defecto si no se encuentra
        
        # Si no tenemos medidas, retornar nulos
        if not alto or not ancho:
            return pd.Series({
                'Z_Sugerido': None,
                'Desarrollo_mm': None,
                'gap_vertical': None,
                'Ancho_rollo_mm': ancho_rollo,
                'n_vertical': None,
                'gap_horizontal_real': None,
                'etq_eje_horizontal': None,
                'Etiquetas_repeticion': None,
                'Repeticiones': None,
                'Merma_mm': None,
                'Merma_m': None,
                'ML_sin_merma': None,
                'M2_sin_merma': None,
                'motivo_imposibilidad': 'SIN MEDIDAS'
            })
        
        # Si Etiquetas_a_Producir es 0, usar 1 como valor minimo para poder calcular Z
        # (asi sabemos que cilindro se necesitaria aunque no haya cantidad pedida)
        unidades_para_calculo = max(1, int(etiquetas_a_producir)) if etiquetas_a_producir > 0 else 1
        
        # Obtener Z óptimo (pasando ancho_rollo extraído)
        resultado_z = obtener_z_sugerido(alto, ancho, unidades_para_calculo, ancho_rollo_mm=ancho_rollo)
        
        if resultado_z is None:
            # Completamente imposible
            return pd.Series({
                'Z_Sugerido': None,
                'Desarrollo_mm': None,
                'gap_vertical': None,
                'Ancho_rollo_mm': ancho_rollo,
                'n_vertical': None,
                'gap_horizontal_real': None,
                'etq_eje_horizontal': None,
                'Etiquetas_repeticion': None,
                'Repeticiones': None,
                'Merma_mm': None,
                'Merma_m': None,
                'ML_sin_merma': None,
                'M2_sin_merma': None,
                'motivo_imposibilidad': f'IMPOSIBLE: Etiqueta {int(alto)}x{int(ancho)}mm no cabe en ningun rollo disponible'
            })
        
        # Verificar si es técnicamente imposible (es_valido=False)
        if resultado_z.get('es_valido') == False:
            # Guardar en lista de impossibilidades
            motivo = resultado_z.get('motivo_imposibilidad', 'Imposibilidad tecnica desconocida')
            imposibilidades_tecnicas.append({
                'Codigo': codigo,
                'Componente': nombre_componente,
                'Alto_mm': int(alto),
                'Ancho_mm': int(ancho),
                'Z_Asignado': resultado_z.get('z'),
                'Rollo_mm': int(resultado_z.get('ancho_rollo_mm', 0)),
                'Motivo': motivo
            })
            
            # Retornar con rollo y Z pero GAP de 0
            return pd.Series({
                'Z_Sugerido': resultado_z.get('z'),
                'Desarrollo_mm': None,
                'gap_vertical': None,
                'Ancho_rollo_mm': resultado_z.get('ancho_rollo_mm'),
                'n_vertical': None,
                'gap_horizontal_real': 0,  # GAP de 0 indica imposibilidad
                'etq_eje_horizontal': None,
                'Etiquetas_repeticion': None,
                'Repeticiones': None,
                'Merma_mm': None,
                'Merma_m': None,
                'ML_sin_merma': None,
                'M2_sin_merma': None,
                'motivo_imposibilidad': motivo
            })
        
        # Cálculo exitoso
        # Calcular merma (ajuste de 0.75 mm por repetición)
        repeticiones = resultado_z.get('repeticiones', 0)
        merma_total_mm = repeticiones * AJUSTE_DESARROLLO_REP_MM
        etiquetas_por_rep = resultado_z.get('n', 0) * resultado_z.get('etq_eje', 0)
        
        # Calcular GAP horizontal real si es None
        gap_horizontal_real = resultado_z.get('gap_horizontal_real')
        if gap_horizontal_real is None and resultado_z.get('etq_eje', 0) == 1:
            # Si cabe solo 1 etiqueta, calcular GAP como espacio a cada lado
            rollo = resultado_z.get('ancho_rollo_mm', 0)
            gap_horizontal_real = (rollo - ancho) / 2.0 if rollo > ancho else 0
        
        return pd.Series({
            'Z_Sugerido': resultado_z.get('z'),
            'Desarrollo_mm': round(resultado_z.get('desarrollo_mm', 0), 2),
            'gap_vertical': round(resultado_z.get('gap_vertical', 0), 2) if resultado_z.get('gap_vertical') is not None else None,
            'Ancho_rollo_mm': resultado_z.get('ancho_rollo_mm'),
            'n_vertical': resultado_z.get('n'),
            'gap_horizontal_real': round(gap_horizontal_real, 2) if gap_horizontal_real is not None else None,
            'etq_eje_horizontal': resultado_z.get('etq_eje'),
            'Etiquetas_repeticion': etiquetas_por_rep,
            'Repeticiones': round(repeticiones, 2) if repeticiones else None,
            'Merma_mm': round(merma_total_mm, 2),
            'Merma_m': round(merma_total_mm / 1000, 4),
            'ML_sin_merma': round(resultado_z.get('ml', 0), 2),
            'M2_sin_merma': round(resultado_z.get('m2', 0), 2),
            'motivo_imposibilidad': ""
        })
    
    # Aplicar cálculo por fila
    calculo_metraje = df_pedidos_componentes_stock.apply(calcular_z_y_metraje, axis=1)
    df_pedidos_componentes_stock = pd.concat([df_pedidos_componentes_stock, calculo_metraje], axis=1)

    # 6. Calcular metros cuadrados pendientes (Etiqueta_m2 × Pendiente)
    df_pedidos_componentes_stock['Etiqueta_m2_pendiente'] = df_pedidos_componentes_stock['Etiqueta_m2'] * df_pedidos_componentes_stock['Pendiente']

    # 7. NOTA: El cascadeo de StockAcumulado se realiza en crear_tabla_resumen() 
    #    NO aquí, para evitar generar valores negativos fantásticos en primeros registros
    #    de componentes sin stock en BD.
    #    El StockAcumulado ya fue calculado correctamente en fillna(0) en línea 957

    # 9. Eliminar columna PendienteAcumulado si existe
    if 'PendienteAcumulado' in df_pedidos_componentes_stock.columns:
        df_pedidos_componentes_stock = df_pedidos_componentes_stock.drop(columns=['PendienteAcumulado'])

    # 10. Renombrar columnas para mayor claridad
    df_pedidos_componentes_stock = df_pedidos_componentes_stock.rename(columns={
        'Etiqueta_Alto': 'Etiqueta_mm_Alto',
        'Etiqueta_Ancho': 'Etiqueta_mm_Ancho',
        'Etiqueta_m2': 'Etiqueta_m2_unitaria',
        'StockFinal': 'StockFinal_m2'
    })

    # 11. Reordenar columnas para visualización intuitiva (incluye motor de cálculo)
    # ═══════════════════════════════════════════════════════════════════════════════
    # ORDEN DE COLUMNAS DIFERENCIADO POR CATEGORÍA
    # ═══════════════════════════════════════════════════════════════════════════════
    orden_cols = [
        # ───────────────────────────────────────────────────────────────────────────
        # IDENTIFICACIÓN Y REFERENCIAS
        # ───────────────────────────────────────────────────────────────────────────
        'FechaEntrega', 'Codigo', 'Prod_padre', 'Componente', 'Nomb_componente',
        
        # ───────────────────────────────────────────────────────────────────────────
        # STOCK Y PENDIENTE
        # ───────────────────────────────────────────────────────────────────────────
        'Pendiente', 'StockAcumulado',
        
        # ───────────────────────────────────────────────────────────────────────────
        # DIMENSIONES DE ETIQUETA (INPUT del motor)
        # ───────────────────────────────────────────────────────────────────────────
        'Etiqueta_mm_Alto', 'Etiqueta_mm_Ancho', 'Ancho_rollo_mm',
        'gap_horizontal_real',      # GAP horizontal (resultado)
        'Etiqueta_m2_unitaria', 'Etiqueta_m2_pendiente',
        
        # ───────────────────────────────────────────────────────────────────────────
        # MOTOR DE CÁLCULO - CILINDRO Y DESARROLLO
        # ───────────────────────────────────────────────────────────────────────────
        'Z_Sugerido',               # Cilindro elegido (60-168)
        'Desarrollo_mm',            # Desarrollo del cilindro (Z × 3.175)
        'gap_vertical',             # GAP vertical en desarrollo
        'n_vertical',               # Etiquetas verticales
        
        # ───────────────────────────────────────────────────────────────────────────
        # MOTOR DE CÁLCULO - ROLLO Y DISTRIBUCIÓN HORIZONTAL
        # ───────────────────────────────────────────────────────────────────────────
        'etq_eje_horizontal',       # Etiquetas horizontales
        'Etiquetas_repeticion',     # Etiquetas por repetición
        
        # ───────────────────────────────────────────────────────────────────────────
        # METRAJE Y MERMA
        # ───────────────────────────────────────────────────────────────────────────
        'Repeticiones',             # Vueltas del cilindro
        'Merma_mm', 'Merma_m',      # Merma (mm y m)
        'ML_sin_merma', 'M2_sin_merma',  # Metros lineales y cuadrados
        
        # ───────────────────────────────────────────────────────────────────────────
        # STOCK FINAL
        # ───────────────────────────────────────────────────────────────────────────
        'StockFinal_m2'             # Stock final después de consumo
    ]
    cols = [c for c in df_pedidos_componentes_stock.columns if c not in orden_cols]
    df_pedidos_componentes_stock = df_pedidos_componentes_stock[[col for col in orden_cols if col in df_pedidos_componentes_stock.columns] + cols]

    print("\n=== MOTOR DE CÁLCULO INTEGRADO ===", flush=True)
    print("Columnas del Motor de Calculo:", flush=True)
    print("  - Z_Sugerido: Cilindro elegido (60-168) para optimizar metraje", flush=True)
    print("  - Desarrollo_mm: Perimetro del cilindro en mm (Z × 3.175)", flush=True)
    print("  - Ancho_rollo_mm: Ancho estandar del rollo (330 mm)", flush=True)
    print("  - n_vertical: Etiquetas que caben verticalmente en el cilindro", flush=True)
    print("  - etq_eje_horizontal: Etiquetas que caben horizontalmente en el rollo", flush=True)
    print("  - Etiquetas_repeticion: Etiquetas por repeticion (n_vertical × etq_eje_horizontal)", flush=True)
    print("  - Repeticiones: Vueltas del cilindro necesarias", flush=True)
    print("  - Merma_mm: Merma total en milimetros (0.75 mm × Repeticiones)", flush=True)
    print("  - Merma_m: Merma total en metros (Merma_mm / 1000)", flush=True)
    print("  - ML_sin_merma: Metros lineales de material (sin contar merma)", flush=True)
    print("  - M2_sin_merma: Metros cuadrados de material (sin contar merma)", flush=True)
    print("=" * 50, flush=True)
    
    # Mostrar resultados del motor de cálculo
    print("\n=== RESULTADOS DEL MOTOR DE CÁLCULO ===", flush=True)
    print(f"Total de registros procesados: {len(df_pedidos_componentes_stock)}", flush=True)
    print("\nPrimeros 15 registros con valores calculados:", flush=True)
    
    # Seleccionar columnas relevantes del motor
    columnas_motor = ['Codigo', 'Componente', 'Pendiente', 'Etiqueta_mm_Alto', 'Etiqueta_mm_Ancho', 
                      'Z_Sugerido', 'Desarrollo_mm', 'Ancho_rollo_mm',
                      'n_vertical', 'etq_eje_horizontal', 'Etiquetas_repeticion',
                      'Repeticiones', 'Merma_mm', 'Merma_m', 'ML_sin_merma', 'M2_sin_merma']
    columnas_existentes = [col for col in columnas_motor if col in df_pedidos_componentes_stock.columns]
    
    resultado_motor = df_pedidos_componentes_stock[columnas_existentes].head(15)
    print(resultado_motor.to_string(index=False), flush=True)
    
    print("\n" + "=" * 50, flush=True)
    print("Estadísticas del Motor:", flush=True)
    print(f"  - Registros con Z calculado: {df_pedidos_componentes_stock['Z_Sugerido'].notna().sum()}", flush=True)
    print(f"  - ML promedio: {df_pedidos_componentes_stock['ML_sin_merma'].mean():.2f} m", flush=True)
    print(f"  - M2 promedio: {df_pedidos_componentes_stock['M2_sin_merma'].mean():.2f} m²", flush=True)
    print(f"  - ML total: {df_pedidos_componentes_stock['ML_sin_merma'].sum():.2f} m", flush=True)
    print(f"  - M2 total: {df_pedidos_componentes_stock['M2_sin_merma'].sum():.2f} m²", flush=True)
    merma_total_m = df_pedidos_componentes_stock['Merma_m'].sum()
    merma_total_mm = df_pedidos_componentes_stock['Merma_mm'].sum()
    print(f"  - Merma total: {merma_total_mm:.2f} mm = {merma_total_m:.2f} m", flush=True)
    print("=" * 50, flush=True)

except Exception as e:
    print(f"Error en la conexión o consulta: {e}", flush=True)
    print(f"\n❌ INSTALACIÓN FALTANTE: Por favor instala pyodbc:", flush=True)
    print(f"   pip install pyodbc", flush=True)
    sys.exit(1)

engine.dispose()

# =====================================================================
# CREAR TABLA RESUMEN VITAL
# =====================================================================
# Crear la tabla resumen con todas las columnas nuevas
df_resumen = crear_tabla_resumen(
    df_pedidos_componentes_stock,
    df_pedidos_pendientes,
    df_listademateriales
)

# Usar RUTA_SALIDA para guardar en la carpeta del ejecutable o Descargas
excel_filename = os.path.join(RUTA_SALIDA, 'tablas_unificadas.xlsx')

# Función para aplicar colores condicionales a columnas de disponibilidad
def aplicar_colores_disponibilidad(workbook_path, sheet_name='TablaResumen'):
    """
    Aplica formato condicional de colores a las columnas Disponibilidad y Disponibilidad_Factor
    1 (verde), 0 (rojo)
    """
    try:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]
        
        # Colores
        verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Verde
        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Rojo
        
        # Fuente blanca para mejor contraste
        fuente_blanca = Font(color="FFFFFF", bold=True)
        
        # Buscar las columnas Disponibilidad y Disponibilidad_Factor
        col_disponibilidad = None
        col_disponibilidad_factor = None
        
        for cell in ws[1]:
            if cell.value == 'Disponibilidad':
                col_disponibilidad = cell.column
            elif cell.value == 'Disponibilidad_Factor':
                col_disponibilidad_factor = cell.column
        
        # Aplicar colores a Disponibilidad
        if col_disponibilidad:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_disponibilidad)
                if cell.value == 1:
                    cell.fill = verde
                    cell.font = fuente_blanca
                elif cell.value == 0:
                    cell.fill = rojo
                    cell.font = fuente_blanca
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar colores a Disponibilidad_Factor
        if col_disponibilidad_factor:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_disponibilidad_factor)
                if cell.value == 1:
                    cell.fill = verde
                    cell.font = fuente_blanca
                elif cell.value == 0:
                    cell.fill = rojo
                    cell.font = fuente_blanca
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(workbook_path)
        print(f"[OK] Colores condicionales aplicados a {sheet_name}", flush=True)
    except Exception as e:
        print(f"[ADVERTENCIA] No se pudieron aplicar colores: {e}", flush=True)

try:
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Tablas principales - Las que usas
        df_stock.to_excel(writer, sheet_name='StockGeneral', index=False)
        df_listademateriales.to_excel(writer, sheet_name='ListaMateriales', index=False)
        df_stock_acumulado.to_excel(writer, sheet_name='StockAcumulado', index=False)
        df_resumen.to_excel(writer, sheet_name='TablaResumen', index=False)
        
        # Tabla técnica - Para auditoría del motor de cálculo
        df_pedidos_componentes_stock.to_excel(writer, sheet_name='PedidosComponentesStock', index=False)
    
    # Aplicar colores condicionales después de guardar
    aplicar_colores_disponibilidad(excel_filename, 'TablaResumen')
    
    print(f"[OK] Archivo guardado: {excel_filename}", flush=True)
except PermissionError:
    excel_filename = os.path.join(RUTA_SALIDA, 'tablas_unificadas_temp.xlsx')
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Tablas principales - Las que usas
        df_stock.to_excel(writer, sheet_name='StockGeneral', index=False)
        df_listademateriales.to_excel(writer, sheet_name='ListaMateriales', index=False)
        df_stock_acumulado.to_excel(writer, sheet_name='StockAcumulado', index=False)
        df_resumen.to_excel(writer, sheet_name='TablaResumen', index=False)
        
        # Tabla técnica - Para auditoría del motor de cálculo
        df_pedidos_componentes_stock.to_excel(writer, sheet_name='PedidosComponentesStock', index=False)
    
    # Aplicar colores condicionales después de guardar
    aplicar_colores_disponibilidad(excel_filename, 'TablaResumen')
    
    print(f"⚠️  Archivo bloqueado, guardado como: {excel_filename}", flush=True)

print("\n[INFO] Proceso completado. Cerrando en 5 segundos...", flush=True)
for i in range(5, 0, -1):
    print(f"[INFO] Cerrando en {i}...", flush=True)
    time.sleep(1)
print("[OK] ¡Listo!", flush=True)
# Ejemplo en PowerShell o cmd:
# python "c:\Users\innjguadalupe\OneDrive - Soluciones de etiquetado Innoprint SA\Escritorio\PYTHON\unificacion.py"